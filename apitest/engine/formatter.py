import json
from pathlib import Path

import yaml

from apitest.models.example import TestExample, TestPlan, TestPlanPhase


def write_examples(examples: list[TestExample], filepath: str, fmt: str) -> None:
    path = Path(filepath)
    if fmt == "json":
        data = {"examples": [e.to_dict() for e in examples]}
        path.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    elif fmt in ("yaml", "yml"):
        data = {"examples": [e.to_dict() for e in examples]}
        path.write_text(yaml.dump(data, allow_unicode=True, sort_keys=False))
    elif fmt == "md":
        _write_examples_md(examples, path)
    elif fmt == "xlsx":
        _write_examples_xlsx(examples, path)
    else:
        raise ValueError(f"Unknown format: {fmt}")


def read_examples(filepath: str, fmt: str) -> list[TestExample]:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Examples file not found: {filepath}")
    if fmt == "json":
        data = json.loads(path.read_text())
    elif fmt in ("yaml", "yml"):
        data = yaml.safe_load(path.read_text())
    elif fmt == "md":
        raise ValueError("Cannot parse examples from markdown. Use json or yaml for machine reading.")
    elif fmt == "xlsx":
        return _read_examples_xlsx(path)
    else:
        raise ValueError(f"Unknown format: {fmt}")
    return [TestExample.from_dict(e) for e in data.get("examples", [])]


def write_plan(plan: TestPlan, filepath: str, fmt: str) -> None:
    path = Path(filepath)
    if fmt == "json":
        path.write_text(json.dumps(plan.to_dict(), indent=2, ensure_ascii=False))
    elif fmt in ("yaml", "yml"):
        path.write_text(yaml.dump(plan.to_dict(), allow_unicode=True, sort_keys=False))
    elif fmt == "md":
        _write_plan_md(plan, path)
    elif fmt == "xlsx":
        _write_plan_xlsx(plan, path)
    else:
        raise ValueError(f"Unknown format: {fmt}")


def read_plan(filepath: str, fmt: str) -> TestPlan:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Plan file not found: {filepath}")
    if fmt == "json":
        data = json.loads(path.read_text())
    elif fmt in ("yaml", "yml"):
        data = yaml.safe_load(path.read_text())
    elif fmt == "md":
        raise ValueError("Cannot parse plan from markdown. Use json or yaml for machine reading.")
    elif fmt == "xlsx":
        return _read_plan_xlsx(path)
    else:
        raise ValueError(f"Unknown format: {fmt}")
    return TestPlan.from_dict(data)


def _write_examples_md(examples: list[TestExample], path: Path) -> None:
    lines = ["# Test Examples\n"]
    for cat in ["happy-path", "equivalence-class", "boundary-value", "negative", "auth-security", "lifecycle"]:
        cat_examples = [e for e in examples if e.category == cat]
        if not cat_examples:
            continue
        lines.append(f"## {cat.replace('-', ' ').title()}\n")
        for e in cat_examples:
            lines.append(f"### {e.id}: {e.description}")
            lines.append(f"- **Endpoint:** `{e.endpoint}`")
            lines.append(f"- **Expected:** {e.expected_status}")
            if e.preconditions:
                lines.append(f"- **Preconditions:** {', '.join(e.preconditions)}")
            if e.depends_on:
                lines.append(f"- **Depends on:** `{e.depends_on}`")
            lines.append("")
    path.write_text("\n".join(lines))


def _write_examples_xlsx(examples: list[TestExample], path: Path) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Examples"
    ws.append(["ID", "Area", "Category", "Endpoint", "Description", "Expected Status",
                "Preconditions", "Depends On", "Cleanup"])
    for e in examples:
        ws.append([e.id, e.area, e.category, e.endpoint, e.description, e.expected_status,
                   ", ".join(e.preconditions), e.depends_on or "", e.cleanup])
    wb.save(path)


def _read_examples_xlsx(path: Path) -> list[TestExample]:
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    examples = []
    for row in rows:
        examples.append(TestExample(
            id=row[0], area=row[1], category=row[2], endpoint=row[3],
            description=row[4], expected_status=row[5],
            preconditions=[p.strip() for p in (row[6] or "").split(",") if p.strip()],
            depends_on=row[7] or None, cleanup=row[8] or "",
        ))
    return examples


def _write_plan_md(plan: TestPlan, path: Path) -> None:
    lines = [f"# {plan.title}\n"]
    lines.append(f"**Coverage:** {plan.coverage} | **Areas:** {', '.join(plan.areas)}")
    lines.append(f"**Total Examples:** {plan.total_examples}\n")
    for phase in plan.phases:
        lines.append(f"## Phase {phase.order}: {phase.name}")
        if phase.description:
            lines.append(f"\n{phase.description}\n")
        if phase.depends_on_phase:
            lines.append(f"\n*Depends on Phase {phase.depends_on_phase}*\n")
        lines.append("**Examples:**")
        for eid in phase.examples:
            lines.append(f"- `{eid}`")
        lines.append("")
    path.write_text("\n".join(lines))


def _write_plan_xlsx(plan: TestPlan, path: Path) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Plan"
    ws.append(["Phase", "Name", "Description", "Depends On Phase", "Example IDs"])
    for phase in plan.phases:
        ws.append([phase.order, phase.name, phase.description,
                   phase.depends_on_phase or "", ", ".join(phase.examples)])
    wb.save(path)


def _read_plan_xlsx(path: Path) -> TestPlan:
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    phases = []
    for row in rows:
        example_ids = [e.strip() for e in (row[4] or "").split(",") if e.strip()]
        phases.append(TestPlanPhase(
            name=row[1], order=row[0],
            description=row[2] or "",
            depends_on_phase=row[3] if row[3] else None,
            examples=example_ids,
        ))
    return TestPlan(title="", coverage="", areas=[], phases=phases)
